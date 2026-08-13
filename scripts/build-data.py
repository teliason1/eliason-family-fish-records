"""Build normalized JSON and web photos from the latest family workbook/TWBX."""
from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "TableauFishRecords - Copy"
XLSX = SOURCE / "FishRecords2b.xlsx"
TWBX = SOURCE / "Family Fish Records.twbx"
DATA_OUT = ROOT / "data" / "records.json"
PHOTO_OUT = ROOT / "public" / "fish"

# Photos present on the thumb drive but not wired into the newest Tableau mapping.
# These names are explicit rather than fuzzy-matched so a catch never receives the
# wrong family photograph.
RECOVERED_PHOTOS = {
    133: "LongfinSquid2.jpg", 119: "RockBass23.jpg", 49: "hake.jpg",
    114: "RainbowTrout6-11-2020.jpg", 54: "hogsucker.jpg", 57: "bonito1.jpg",
    59: "skipjack.jpg", 112: "yellowfin17.jpg", 8: "mahi-mahi17.jpg",
    64: "goatfish.jpg", 65: "porcupine.jpg", 68: "whiteperch.jpg",
    70: "bowfin.jpg", 77: "sheepshead.jpg", 80: "herring11.jpg",
    82: "sculpin83.jpg", 38: "hickshad10.jpg", 5: "BlackDrum.jpg",
    41: "Black Crappie.jpg", 6: "striper08.jpg", 86: "spinydogfish.jpg",
    88: "ray.jpg", 87: "flounder.jpg", 89: "Bunker.jpg", 90: "warmou2.jpg",
    98: "warmou1.jpg", 101: "bullhead.jpg", 33: "FreshwaterDrum.GIF",
    105: "perch.gif", 106: "TrumpetFish.GIF", 16: "BlueFish.GIF",
    27: "SilverSalmon.GIF", 108: "LakeTrout.GIF", 107: "Grayling.GIF",
    36: "DollyVarden.GIF", 24: "ChumSalmon.GIF", 12: "Halibut.GIF",
    28: "PinkSalmon.GIF", 32: "BrownTrout.gif", 7: "ChinookSalmon.GIF",
    110: "CutthroatTrout.GIF", 21: "Pike.GIF",
}

ESTIMATED_COORDINATES = {
    (None, "Alaska"): (64.2008, -149.4937),
    (None, "Maine"): (45.2538, -69.4455),
    (None, "Massachusetts"): (42.4072, -71.3824),
    ("Bar Harbor", "Maine"): (44.3876378, -68.2043361),
    ("Byers Creek", "Alaska"): (62.7105922, -150.2015929),
    ("Corpus Christi", "Texas"): (27.7635302, -97.4033191),
    ("Gloucester", "Massachusetts"): (42.6153595, -70.6624608),
    ("Greencastle", "Indiana"): (39.6444898, -86.8647316),
    ("Homer", "Alaska"): (59.6454064, -151.5445643),
    ("Hot Springs", "Arkansas"): (34.5038393, -93.0552437),
    ("Indianapolis", "Indiana"): (39.7683331, -86.1583502),
    ("Kennebunkport", "Maine"): (43.3614989, -70.4771132),
    ("Marblehead", "Massachusetts"): (42.5000960, -70.8578253),
    ("Richmond", "Virginia"): (37.5385087, -77.4342800),
}


def slug(value: str) -> str:
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", value.lower()))


def clean(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def mapped_photos(archive: ZipFile) -> tuple[dict[int, str], dict[str, str]]:
    twb = next(name for name in archive.namelist() if name.lower().endswith(".twb"))
    root = ET.fromstring(archive.read(twb))
    id_mapping = {}
    species_mapping = {}
    for image in root.findall(".//mapped-image"):
        expression = image.get("expression", "")
        if not expression.startswith("file:Image/"):
            continue
        id_filter = image.find(".//groupfilter[@level='[ID]']")
        species_filter = image.find(".//groupfilter[@level='[Fish Species]']")
        if id_filter is not None:
            id_mapping[int(id_filter.get("member"))] = expression.removeprefix("file:")
        elif species_filter is not None:
            species_mapping[species_filter.get("member").strip('"')] = expression.removeprefix("file:")
    return id_mapping, species_mapping


def main():
    DATA_OUT.parent.mkdir(parents=True, exist_ok=True)
    PHOTO_OUT.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(XLSX, read_only=True, data_only=True)
    records = []
    with ZipFile(TWBX) as archive:
        photos, species_photos = mapped_photos(archive)
        for sheet_name in ("Index Sheet",):
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows())]
            for values in sheet.iter_rows(min_row=2, values_only=True):
                if values[0] is None:
                    continue
                row = {key: clean(value) for key, value in zip(headers, values)}
                record_id = int(row["ID"])
                photo_url = None
                if record_id in photos or record_id in RECOVERED_PHOTOS or row["Fish Species"] in species_photos:
                    archive_name = photos.get(record_id) or (f"Image/{RECOVERED_PHOTOS[record_id]}" if record_id in RECOVERED_PHOTOS else species_photos[row["Fish Species"]])
                    suffix = Path(archive_name).suffix.lower().replace(".gif", ".gif")
                    filename = f"{record_id}-{slug(row['Fish Species'])}{suffix}"
                    destination = PHOTO_OUT / filename
                    if archive_name in archive.namelist():
                        with archive.open(archive_name) as src, destination.open("wb") as dst:
                            shutil.copyfileobj(src, dst)
                    else:
                        shutil.copy2(SOURCE / Path(archive_name).name, destination)
                    photo_url = f"/fish/{filename}"
                estimate_key = (row["City"], row["State"]) if (row["City"], row["State"]) in ESTIMATED_COORDINATES else (None, row["State"])
                estimated = row["Lat"] is None and estimate_key in ESTIMATED_COORDINATES
                lat, lng = (ESTIMATED_COORDINATES[estimate_key] if estimated else (row["Lat"], row["Long"]))
                records.append({
                    "id": record_id,
                    "species": row["Fish Species"],
                    "angler": row["Angler"],
                    "date": row["Date"],
                    "weight": float(row["Weight"]) if row["Weight"] is not None else None,
                    "length": float(row["Length"]) if row["Length"] is not None else None,
                    "state": row["State"],
                    "city": row["City"],
                    "water": row["Water"],
                    "caughtWith": row["Caught With"],
                    "status": {"Y": "current", "N": "historical", "M": "micro"}[row["Current"]],
                    "story": row["Story"],
                    "lat": float(lat) if lat is not None else None,
                    "lng": float(lng) if lng is not None else None,
                    "coordinateAccuracy": "estimated" if estimated else "exact" if lat is not None else "unknown",
                    "photo": photo_url,
                })
    records.sort(key=lambda item: (item["date"], item["id"]), reverse=True)
    DATA_OUT.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n")
    print(f"Generated {len(records)} records and {sum(bool(r['photo']) for r in records)} mapped photos")


if __name__ == "__main__":
    main()
